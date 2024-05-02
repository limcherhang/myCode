import os
import sys
rootPath = os.getcwd() + '/../'
sys.path.append(rootPath)
import configparser
from connection.mongo_connection import MongoConn
from bson import ObjectId
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read(rootPath+'/config.ini')


    # Create mongo connection
    client = MongoConn(config['mongo_production_nxmap'])
    client.connect()

    # Get database
    db = client.get_database()

    companyName = sys.argv[1].replace('_', ' ')

    users_info = db.users.find({"companyName": companyName})

    user = users_info[0]

    companyId = user['companyId']
    # print(companyId)
    asset_infos = {}
    
    company_asset = db.company_assets.find({"company": ObjectId(companyId)}, {"companyAsset": 1, "_id": 0})[0]['companyAsset']  # Get company asset
    for ca in company_asset:
        
        for asset in ca['assets']:
            
            asset_infos[asset['_id']] = {
                "Name": asset.get('assetName', asset.get('name', asset.get('transportationName', asset.get('supplierName')))),
                "emissionSource": asset.get('emissionSource', asset.get('activityType')),
                "biomass": "No",
                "scope": ca.get('scope', None),
                "category": asset.get('category', None),
                "emissionType": asset.get('fuelType', asset.get('calculateType', asset.get('calculationApproach'))),
                "unit": asset.get('unit', None),
                "co2EmissionFactor": asset.get("baseEmissionFactor", None),
                "co2unit": asset.get("emissionUnit", None),
                "ch4EmissionFactor": asset.get("ch4EmissionValue", None),
                "ch4unit": asset.get("ch4Unit"),
                "n2oEmissonFactor": asset.get("n2oEmissionValue", None),
                "n2ounit": asset.get("n2oUnit")
            }

    data = db.assetdatas.find({"company": ObjectId(companyId)})[0]  # get company asset data

    consumption_data = data['consumption_data']

    emission_data = data['emission_data']

    consumptionValue = {}
    emissionValue = {}

    for cd in consumption_data:
        year = cd['year']
        for consumption_yearly_data in cd['consumption_yearly_data']:
            if consumption_yearly_data['assetId'] not in consumptionValue:
                consumptionValue[consumption_yearly_data['assetId']] = {}
            for cv in consumption_yearly_data['consumptionValue']:
                consumptionValue[consumption_yearly_data['assetId']][cv['_id']] = {
                    "year": year,
                    "consumption_value": cv['consumptionValue'],
                    "co2GWP": 1,
                    "ch4GWP": 28,       
                    "n2oGWP": 265
                }       # collect consumption yearly data

    for ed in emission_data:
        for emission_yearly_data in ed['emission_yearly_data']:
            if emission_yearly_data['assetId'] not in emissionValue:
                emissionValue[emission_yearly_data['assetId']] = {}

            for ev in emission_yearly_data['emissionDataValue']:
                emissionValue[emission_yearly_data['assetId']][ev['consumptionId']] = {
                    "co2EmissionValue": ev['emissionValue'],
                    "ch4EmissionValue": ev['ch4EmissionValue'],
                    "n2oEmissonValue": ev['n2oEmissonValue'],
                    "co2GWPEmissionValue": ev['co2GWPEmissionValue'],
                    "ch4GWPEmissionValue": ev['ch4GWPEmissionValue'],
                    "n2oGWPEmissionValue": ev['n2oGWPEmissionValue']
                }       # collect emission data value

    ans = []
    i = 1
    total4ratio = 0
    for key, cv in consumptionValue.items():
        
        ev = emissionValue[key]
        ai = asset_infos[key]

        for sub_k, sub_cv in cv.items():

            sub_ev = ev[sub_k]

            activityData = sub_cv['consumption_value']
            
            co2EmissionFactor = ai["co2EmissionFactor"]
            co2EmissionValue = round(activityData * co2EmissionFactor, 2)
            co2GWP = sub_cv['co2GWP']
            co2GWPEmissionValue = round(co2EmissionValue*co2GWP, 2)

            ch4EmissionFactor = ai["ch4EmissionFactor"] if ai["ch4EmissionFactor"] is not None else 0
            ch4EmissionValue = round(activityData * ch4EmissionFactor, 2)
            ch4GWP = sub_cv['ch4GWP']
            ch4GWPEmissionValue = round(ch4EmissionValue*ch4GWP, 2)

            n2oEmissonFactor = ai["n2oEmissonFactor"] if ai["n2oEmissonFactor"] is not None else 0
            n2oEmissionValue = round(activityData * n2oEmissonFactor, 2)
            n2oGWP = sub_cv['n2oGWP']
            n2oGWPEmissionValue = round(n2oEmissionValue*n2oGWP, 2)
            
            subtotal_of_emission_equivalent_from_a_single_emission_source = co2GWPEmissionValue + ch4GWPEmissionValue + n2oGWPEmissionValue     # subtotal of emission equivalent from a single emission source

            total4ratio += subtotal_of_emission_equivalent_from_a_single_emission_source        # sum of all total


            # append all asset data to a list
            ans.append([
                i,                                                                      # No
                ai["Name"],                                                             # Name
                ai['emissionSource'],                                                   # Emission Source
                ai['biomass'],                                                          # Biomass
                ai['scope'],                                                            # Scope
                ai['category'],                                                         # Category
                ai['emissionType'],                                                     # Emission Type
                activityData,                                                           # Activity Data(Year) 
                ai['unit'],                                                             # Unit
                "CO2",                                                                  # GHG#1
                co2EmissionFactor,                                                      # Emission Factor
                ai['co2unit'],                                                          # Unit
                co2EmissionValue,                                                       # Emission (tonne/Year)
                co2GWP,                                                                 # GWP
                co2GWPEmissionValue,                                                    # Emission equivalent (tCO2e/Year)
                "CH4",                                                                  # GHG#2
                ch4EmissionFactor,                                                      # Emission Factor
                ai['ch4unit'],                                                          # Unit
                ch4EmissionValue,                                                       # Emission (tonne/Year)
                ch4GWP,                                                                 # GWP
                ch4GWPEmissionValue,                                                    # Emission equivalent(tCO2e/Year)
                "N2O",                                                                  # GHG#3
                n2oEmissonFactor,                                                       # Emission Factor
                ai['n2ounit'],                                                          # Unit
                n2oEmissionValue,                                                       # Emission (tonne/Year)
                n2oGWP,                                                                 # GWP
                n2oGWPEmissionValue,                                                    # Emission equivalent(tCO2e/Year)
                subtotal_of_emission_equivalent_from_a_single_emission_source,          # Subtotal of emission equivalents from a single emission source(tCO2e/Year)
                co2GWPEmissionValue if ai['biomass'] == 'yes' else 0,                   # Subtotal of CO2 emission equivalent of biomass fuel from a single emission source (tCO2e/Year)
                subtotal_of_emission_equivalent_from_a_single_emission_source           # Ratio of single emission source to total emissions (%)
            ])

            i+=1
    
    # tranform list to pandas dataframe
    df = pd.DataFrame(ans)
    df.iloc[:, -1] = df.iloc[:, -1] / total4ratio * 100         # complete the ratio
    df.iloc[:, -1] = df.iloc[:, -1].round(3)
    df.iloc[:, -1] = df.iloc[:, -1].apply(lambda x: f"{x}%")

    # Processing dataframe to excel file
    with pd.ExcelWriter(f"{companyName} GHG Emission Information.xlsx", engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="English Version", index=False, startcol=0, startrow=2, header=False)

        worksheet = writer.sheets['English Version']

        # merge cell A1 & A2
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.cell(row=1, column=1, value='No').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell B1 & B2
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        worksheet.cell(row=1, column=2, value='Name').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell C1 & C2
        worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        worksheet.cell(row=1, column=3, value='Emission Source').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell D1 & D2
        worksheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        worksheet.cell(row=1, column=4, value='Biomass').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell E1 & E2
        worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)
        worksheet.cell(row=1, column=5, value='Scope').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell F1 & F2
        worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
        worksheet.cell(row=1, column=6, value='Category').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell G1 & G2
        worksheet.merge_cells(start_row=1, start_column=7, end_row=2, end_column=7)
        worksheet.cell(row=1, column=7, value='Emission Type').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell H1 & H2
        worksheet.merge_cells(start_row=1, start_column=8, end_row=2, end_column=8)
        worksheet.cell(row=1, column=8, value='Activity Data(Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell I1 & I2
        worksheet.merge_cells(start_row=1, start_column=9, end_row=2, end_column=9)
        worksheet.cell(row=1, column=9, value='Unit').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell J1 to O1
        worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=15)

        # Alignment J2 to O2
        worksheet.cell(row=1, column=10, value=r'GHG#1').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=10, value=r'GHG#1').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=11, value='Emission Factor').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=12, value='Unit').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=13, value='Emission (tonne/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=14, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=15, value='Emission equivalent\n(tCO2e/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell P1 to U1
        worksheet.merge_cells(start_row=1, start_column=16, end_row=1, end_column=21)

        # Alignment P2 to U2
        worksheet.cell(row=1, column=16, value=r'GHG#2').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=16, value=r'GHG#2').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=17, value='Emission Factor').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=18, value='Unit').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=19, value='Emission (tonne/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=20, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=21, value='Emission equivalent\n(tCO2e/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell V1 to AA1
        worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=27)

        # Alignment V2 to AA2
        worksheet.cell(row=1, column=22, value=r'GHG#3').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=22, value=r'GHG#3').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=23, value='Emission Factor').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=24, value='Unit').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=25, value='Emission (tonne/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=26, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=27, value='Emission equivalent\n(tCO2e/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell AB1 & AB2 & Alignment
        worksheet.merge_cells(start_row=1, start_column=28,  end_row=2, end_column=28)
        worksheet.cell(row=1, column=28, value='Subtotal of\nemission equivalents\nfrom a single emission source\n(tCO2e/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell AC1 & AC2 & Alignment
        worksheet.merge_cells(start_row=1, start_column=29,  end_row=2, end_column=29)
        worksheet.cell(row=1, column=29, value='Subtotal of\nCO2 emission equivalent\nof biomass fuel\nfrom a single emission source\n(tCO2e/Year)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell AD1 & AD2 & Alignment
        worksheet.merge_cells(start_row=1, start_column=30,  end_row=2, end_column=30)
        worksheet.cell(row=1, column=30, value='Ratio of\nsingle emission source\nto\ntotal emissions\n(%)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # custom width
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 20
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
        worksheet.column_dimensions['M'].width = 20
        worksheet.column_dimensions['N'].width = 15
        worksheet.column_dimensions['O'].width = 15
        worksheet.column_dimensions['P'].width = 15
        worksheet.column_dimensions['Q'].width = 15
        worksheet.column_dimensions['R'].width = 15
        worksheet.column_dimensions['S'].width = 20
        worksheet.column_dimensions['T'].width = 15
        worksheet.column_dimensions['U'].width = 15
        worksheet.column_dimensions['V'].width = 15
        worksheet.column_dimensions['W'].width = 15
        worksheet.column_dimensions['X'].width = 15
        worksheet.column_dimensions['Y'].width = 20
        worksheet.column_dimensions['Z'].width = 15
        worksheet.column_dimensions['AA'].width = 15
        worksheet.column_dimensions['AB'].width = 20
        worksheet.column_dimensions['AC'].width = 20
        worksheet.column_dimensions['AD'].width = 20

        # custom height
        worksheet.row_dimensions[1].height = 55
        worksheet.row_dimensions[2].height = 55  

        # adjust the final width
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max(
                series.astype(str).map(len).max(),
                len(str(series.name))
            )+5

            current_column_dimension = worksheet.column_dimensions[get_column_letter(idx+1)]
            current_width = current_column_dimension.width

            max_len = max(max_len, current_width)

            worksheet.column_dimensions[get_column_letter(idx+1)].width = max_len

        # Chinese Version
        df.to_excel(writer, sheet_name="Chinese Version", index=False, startcol=0, startrow=2, header=False)

        worksheet = writer.sheets['Chinese Version']

        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.cell(row=1, column=1, value='編號').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        worksheet.cell(row=1, column=2, value='製程名稱').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        worksheet.cell(row=1, column=3, value='排放源').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
        worksheet.cell(row=1, column=4, value='是否屬於生質能源').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=5)
        worksheet.cell(row=1, column=5, value='範疇').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=6, end_row=2, end_column=6)
        worksheet.cell(row=1, column=6, value='類別').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=7, end_row=2, end_column=7)
        worksheet.cell(row=1, column=7, value='排放類型').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=8, end_row=2, end_column=8)
        worksheet.cell(row=1, column=8, value='年活動數據').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=9, end_row=2, end_column=9)
        worksheet.cell(row=1, column=9, value='單位').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=15)
        worksheet.cell(row=1, column=10, value=r'溫室氣體#1').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=10, value=r'溫室氣體#1').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=11, value='排放係數').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=12, value='係數單位').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=13, value='排放量\n(公噸/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=14, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=15, value='排放當量\n(公噸CO2e/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=16, end_row=1, end_column=21)
        worksheet.cell(row=1, column=16, value=r'溫室氣體#2').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=16, value=r'溫室氣體#2').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=17, value='排放係數').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=18, value='係數單位').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=19, value='排放量\n(公噸/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=20, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=21, value='排放當量\n(公噸CO2e/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=27)
        worksheet.cell(row=1, column=22, value=r'溫室氣體#3').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=22, value=r'溫室氣體#3').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=23, value='排放係數').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=24, value='係數單位').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=25, value='排放量\n(公噸/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=26, value='GWP').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=27, value='排放當量\n(公噸CO2e/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=28,  end_row=2, end_column=28)
        worksheet.cell(row=1, column=28, value='單一排放源排放當量小計\n(CO2e公噸/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=29,  end_row=2, end_column=29)
        worksheet.cell(row=1, column=29, value='單一排放源生質燃料CO2排放當量小計\n(CO2e公噸/年)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=30,  end_row=2, end_column=30)
        worksheet.cell(row=1, column=30, value='單一排放源占排放總量比\n(%)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 20
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
        worksheet.column_dimensions['M'].width = 20
        worksheet.column_dimensions['N'].width = 15
        worksheet.column_dimensions['O'].width = 15
        worksheet.column_dimensions['P'].width = 15
        worksheet.column_dimensions['Q'].width = 15
        worksheet.column_dimensions['R'].width = 15
        worksheet.column_dimensions['S'].width = 20
        worksheet.column_dimensions['T'].width = 15
        worksheet.column_dimensions['U'].width = 15
        worksheet.column_dimensions['V'].width = 15
        worksheet.column_dimensions['W'].width = 15
        worksheet.column_dimensions['X'].width = 15
        worksheet.column_dimensions['Y'].width = 20
        worksheet.column_dimensions['Z'].width = 15
        worksheet.column_dimensions['AA'].width = 15
        worksheet.column_dimensions['AB'].width = 20
        worksheet.column_dimensions['AC'].width = 20
        worksheet.column_dimensions['AD'].width = 20

        worksheet.row_dimensions[1].height = 55  
        worksheet.row_dimensions[2].height = 55
        
        for idx, col in enumerate(df):
            series = df[col]
            max_len = max(
                series.astype(str).map(len).max(),
                len(str(series.name))
            )+5

            current_column_dimension = worksheet.column_dimensions[get_column_letter(idx+1)]
            current_width = current_column_dimension.width

            max_len = max(max_len, current_width)

            worksheet.column_dimensions[get_column_letter(idx+1)].width = max_len