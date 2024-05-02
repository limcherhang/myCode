import os
import sys
rootPath = os.getcwd() + '/../'
sys.path.append(rootPath)
import configparser
from connection.mongo_connection import MongoConn
from bson import ObjectId
import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Protection
from openpyxl.worksheet.datavalidation import DataValidation

if __name__ == '__main__':
    config = configparser.ConfigParser()
    config.read(rootPath+'/config.ini')


    # Create mongo connection
    client = MongoConn(config['mongo_production_nxmap'])
    client.connect()

    # Get database
    db = client.get_database()

    companyName = sys.argv[1].replace('_', ' ')

    users_info = db.users.find({"companyName": "Taiwan Aaron"})

    user = users_info[0]

    companyId = user['companyId']
    # print(companyId)
    asset_infos = {}
    
    company_asset = db.company_assets.find({"company": ObjectId(companyId)}, {"companyAsset": 1, "_id": 0})[0]['companyAsset']
    for ca in company_asset:
        
        for asset in ca['assets']:
            
            asset_infos[asset['_id']] = {
                "Name": asset.get('assetName', asset.get('name', asset.get('transportationName', asset.get('supplierName')))),
                "emissionSource": asset.get('emissionSource', asset.get('activityType')),
                "biomass": "No",
                "scope": ca.get('scope', None),
                "category": asset.get('category', None),
                "emissionType": asset.get('fuelType', asset.get('calculateType', asset.get('calculationApproach'))),
                "unit": asset.get('unit', None),
                "co2EmissionFactor": asset.get("baseEmissionFactor", None),
                "co2unit": asset.get("emissionUnit", None),
                "ch4EmissionFactor": asset.get("ch4EmissionValue", None),
                "ch4unit": asset.get("ch4Unit"),
                "n2oEmissonFactor": asset.get("n2oEmissionValue", None),
                "n2ounit": asset.get("n2oUnit")
            }
    # for k,v in asset_infos.items():
    #     print(k, v)
    #     print()
    data = db.assetdatas.find({"company": ObjectId(companyId)})[0]

    consumption_data = data['consumption_data']

    emission_data = data['emission_data']

    consumptionValue = {}
    emissionValue = {}

    for cd in consumption_data:
        year = cd['year']
        for consumption_yearly_data in cd['consumption_yearly_data']:
            if consumption_yearly_data['assetId'] not in consumptionValue:
                consumptionValue[consumption_yearly_data['assetId']] = {}
            for cv in consumption_yearly_data['consumptionValue']:
                consumptionValue[consumption_yearly_data['assetId']][cv['_id']] = {
                    "year": year,
                    "consumption_value": cv['consumptionValue'],
                    "co2GWP": 1,
                    "ch4GWP": 28,
                    "n2oGWP": 265
                }       # collect consumption yearly data

    for ed in emission_data:
        for emission_yearly_data in ed['emission_yearly_data']:
            if emission_yearly_data['assetId'] not in emissionValue:
                emissionValue[emission_yearly_data['assetId']] = {}

            for ev in emission_yearly_data['emissionDataValue']:
                emissionValue[emission_yearly_data['assetId']][ev['consumptionId']] = {
                    "co2EmissionValue": ev['emissionValue'],
                    "ch4EmissionValue": ev['ch4EmissionValue'],
                    "n2oEmissonValue": ev['n2oEmissonValue'],
                    "co2GWPEmissionValue": ev['co2GWPEmissionValue'],
                    "ch4GWPEmissionValue": ev['ch4GWPEmissionValue'],
                    "n2oGWPEmissionValue": ev['n2oGWPEmissionValue']
                }       # collect emission data value
    
    ans = []
    i = 1
    total4ratio = 0

    for key, cv in consumptionValue.items():

        ev = emissionValue[key]
        ai = asset_infos[key]

        for sub_k, sub_cv in cv.items():

            sub_ev = ev[sub_k]

            activityData = sub_cv['consumption_value']
            
            co2EmissionFactor = ai["co2EmissionFactor"]
            co2EmissionValue = round(activityData * co2EmissionFactor, 2)
            co2GWP = sub_cv['co2GWP']
            co2GWPEmissionValue = round(co2EmissionValue*co2GWP, 2)

            ch4EmissionFactor = ai["ch4EmissionFactor"] if ai["ch4EmissionFactor"] is not None else 0
            ch4EmissionValue = round(activityData * ch4EmissionFactor, 2)
            ch4GWP = sub_cv['ch4GWP']
            ch4GWPEmissionValue = round(ch4EmissionValue*ch4GWP, 2)

            n2oEmissonFactor = ai["n2oEmissonFactor"] if ai["n2oEmissonFactor"] is not None else 0
            n2oEmissionValue = round(activityData * n2oEmissonFactor, 2)
            n2oGWP = sub_cv['n2oGWP']
            n2oGWPEmissionValue = round(n2oEmissionValue*n2oGWP, 2)
            
            subtotal_of_emission_equivalent_from_a_single_emission_source = co2GWPEmissionValue + ch4GWPEmissionValue + n2oGWPEmissionValue

            total4ratio += subtotal_of_emission_equivalent_from_a_single_emission_source


            ans.append([
                i,
                ai['Name'],
                ai['emissionSource'],
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                subtotal_of_emission_equivalent_from_a_single_emission_source
            ])

            i+=1
    

    df = pd.DataFrame(ans)
    df.iloc[:, -1] = df.iloc[:, -1] / total4ratio * 100             # complete the ratio
    df.iloc[:, -1] = df.iloc[:, -1].round(3)
    df.iloc[:, -1] = df.iloc[:, -1].apply(lambda x: f"{x}%")

    filename = f"{companyName} Data Quality.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="English Version", index=False, startcol=0, startrow=2, header=False)

        n_rows = df.shape[0]

        worksheet = writer.sheets['English Version']
        workbook = writer.book

        # merge cell A1 & A2
        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.cell(row=1, column=1, value='No').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell B1 & B2
        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        worksheet.cell(row=1, column=2, value='Name').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell C1 & C2
        worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        worksheet.cell(row=1, column=3, value='Emission Type').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell D1 to G1
        worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)

        # Alignment D2 to G2
        worksheet.cell(row=1, column=4, value='Raw fuel materials or products').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=4, value='Activity Data Type').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=5, value='Activity Data\nType Level').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=6, value='Activity data trusted types\n(Instrument calibration error level)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=7, value='Activity data trusted Level').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell H1 to I1
        worksheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)

        # Alignment H2 to I2
        worksheet.cell(row=1, column=8, value='Emission Factor').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=8, value='Emission Factor Type').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=9, value='Emission Factor Type Level').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # merge cell J1 to M1
        worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)

        # Alignment J2 to M2
        worksheet.cell(row=1, column=10, value='Raw fuel materials or products').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=10, value='Single emission source data error level').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=11, value='Proportion of single emission source\nin total emissions (%)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=12, value='Score range').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=13, value='Weighted average of emissions share').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
        worksheet.column_dimensions['M'].width = 15  

        unlocked_style = NamedStyle(name='unlocked', protection=Protection(locked=False))
        workbook.add_named_style(unlocked_style)

        # Set the first cell style
        worksheet['A1'].style = 'unlocked'

        # Set the style of the entire column
        for row in worksheet.iter_rows(min_row=1, max_row=n_rows, min_col=2, max_col=2):
            for cell in row:
                cell.style = 'unlocked'
            
        # Create drop-down list
        for i in range(3, n_rows + 3):
            dv = DataValidation(type="list", formula1='"Continuous measurement,Periodic (intermittent) measurement,Financial accounting estimates,Self-assessment"', allow_blank=True)
            worksheet.add_data_validation(dv)
            dv.add(f'D{i}')

            dv2 = DataValidation(type="list", formula1='"Those who have performed external calibration or have multiple sets of data to support this,Those with certificates such as internal correction or accounting visa,Failure to perform instrument calibration or record compilation"', allow_blank=True)
            worksheet.add_data_validation(dv2)
            dv2.add(f'F{i}')

            dv3 = DataValidation(type="list", formula1='"In-house development coefficient/mass balance coefficient,Same process/equipment experience coefficient,The manufacturer provides coefficients,Regional emission coefficient,National emission coefficient,International emission coefficient"', allow_blank=True)
            worksheet.add_data_validation(dv3)
            dv3.add(f'H{i}')

            for col in ['D', 'F', 'H']:
                cell = worksheet[f'{col}{i}']
                cell.style = 'locked'

            current_row = i

            formula = f'=IF(D{current_row}="Continuous measurement", 1, IF(D{current_row}="Periodic (intermittent) measurement", 2, IF(D{current_row}="Financial accounting estimates", 3, IF(D{current_row}="Self-assessment", 3, ""))))'
            cell = worksheet.cell(row=current_row, column=5)
            cell.value = formula

            formula = f'=IF(F{current_row}="Those who have performed external calibration or have multiple sets of data to support this", 1, IF(F{current_row}="Those with certificates such as internal correction or accounting visa", 2, IF(F{current_row}="Failure to perform instrument calibration or record compilation", 3, "")))'
            cell = worksheet.cell(row=current_row, column=7)
            cell.value = formula

            formula = f'=IF(H{current_row}="In-house development coefficient/mass balance coefficient", 1, IF(H{current_row}="Same process/equipment experience coefficient", 1, IF(H{current_row}="The manufacturer provides coefficients", 2, IF(H{current_row}="Regional emission coefficient",2, IF(H{current_row}="National emission coefficient", 3, IF(H{current_row}="International emission coefficient", 3, ""))))))'
            cell = worksheet.cell(row=current_row, column=9)
            cell.value = formula

            formula = f'=IF(OR(E{current_row}="", G{current_row}="", I{current_row}=""), "", E{current_row}*G{current_row}*I{current_row})'
            cell = worksheet.cell(row=current_row, column=10)
            cell.value = formula

            formula = f'=IF(J{current_row}="", "", IF(J{current_row}<10, 1, IF(AND(J{current_row}>=10, J{current_row}<19), 2, IF(AND(J{current_row}>=19, J{current_row}<=27), 3, ""))))'
            cell = worksheet.cell(row=current_row, column=12)
            cell.value = formula

            formula = f'=IF(OR(J{current_row}="", K{current_row}=""), "",J{current_row}*K{current_row})'
            cell = worksheet.cell(row=current_row, column=13)
            cell.value = formula

        for idx, col in enumerate(df):
            if idx >= 3:break
            series = df[col]
            max_len = max(
                series.astype(str).map(len).max(),
                len(str(series.name))
            )+5

            current_column_dimension = worksheet.column_dimensions[get_column_letter(idx+1)]
            current_width = current_column_dimension.width

            max_len = max(max_len, current_width)

            worksheet.column_dimensions[get_column_letter(idx+1)].width = max_len


        # Chinese version
        df.to_excel(writer, sheet_name="Chinese Version", index=False, startcol=0, startrow=2, header=False)

        worksheet = writer.sheets['Chinese Version']

        worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        worksheet.cell(row=1, column=1, value='編號').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
        worksheet.cell(row=1, column=2, value='製程名稱').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)
        worksheet.cell(row=1, column=3, value='排放源').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=7)
        worksheet.cell(row=1, column=4, value='原燃物料或產品').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=4, value='活動數據種類').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=5, value='活動數據種類等級').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=6, value='活動數據可信種類\n(儀器校正誤差等級)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=7, value='活動數據可信等級').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=9)
        worksheet.cell(row=1, column=8, value='排放係數').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=8, value='排放係數種類').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=9, value='係數種類等級').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
        worksheet.cell(row=1, column=10, value='原燃物料或產品').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=10, value='單一排放源數據誤差等級').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=11, value='單一排放源占排放總量比(%)').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=12, value='評分區間範圍').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.cell(row=2, column=13, value='排放量佔比加權平均').alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 15
        worksheet.column_dimensions['C'].width = 15
        worksheet.column_dimensions['D'].width = 15
        worksheet.column_dimensions['E'].width = 15
        worksheet.column_dimensions['F'].width = 15
        worksheet.column_dimensions['G'].width = 15
        worksheet.column_dimensions['H'].width = 15
        worksheet.column_dimensions['I'].width = 15
        worksheet.column_dimensions['J'].width = 15
        worksheet.column_dimensions['K'].width = 15
        worksheet.column_dimensions['L'].width = 15
        worksheet.column_dimensions['M'].width = 15

        # Set the first cell style
        worksheet['A1'].style = 'unlocked'

        # Set the style of the entire column
        for row in worksheet.iter_rows(min_row=1, max_row=n_rows, min_col=2, max_col=2):
            for cell in row:
                cell.style = 'unlocked'
            
        # Create drop-down list
        for i in range(3, n_rows + 3):
            dv = DataValidation(type="list", formula1='"連續量測,定期(間歇)量測,財務會計推估,自行評估"', allow_blank=True)
            worksheet.add_data_validation(dv)
            dv.add(f'D{i}')

            dv2 = DataValidation(type="list", formula1='"有進行外部校正或有多組數據茲佐證者,有進行內部校正或經過會計簽證等証明者,未進行儀器校正或未進行紀錄彙整者"', allow_blank=True)
            worksheet.add_data_validation(dv2)
            dv2.add(f'F{i}')

            dv2 = DataValidation(type="list", formula1='"自廠發展係數/質量平衡所得係數,同製程/設備經驗係數,製造廠提供係數,區域排放係數,國家排放係數,國際排放係數"', allow_blank=True)
            worksheet.add_data_validation(dv2)
            dv2.add(f'H{i}')

        for i in range(n_rows):
            current_row = i+3

            formula = f'=IF(D{current_row}="連續量測", 1, IF(D{current_row}="定期(間歇)量測", 2, IF(D{current_row}="財務會計推估", 3, IF(D{current_row}="自行評估", 3, ""))))'
            cell = worksheet.cell(row=current_row, column=5)
            cell.value = formula

            formula = f'=IF(F{current_row}="T有進行外部校正或有多組數據茲佐證者", 1, IF(F{current_row}="有進行內部校正或經過會計簽證等証明者", 2, IF(F{current_row}="未進行儀器校正或未進行紀錄彙整者", 3, "")))'
            cell = worksheet.cell(row=current_row, column=7)
            cell.value = formula

            formula = f'=IF(H{current_row}="自廠發展係數/質量平衡所得係數", 1, IF(H{current_row}="同製程/設備經驗係數", 1, IF(H{current_row}="製造廠提供係數", 2, IF(H{current_row}="區域排放係數",2, IF(H{current_row}="國家排放係數", 3, IF(H{current_row}="國際排放係數", 3, ""))))))'
            cell = worksheet.cell(row=current_row, column=9)
            cell.value = formula

            formula = f'=IF(OR(E{current_row}="", G{current_row}="", I{current_row}=""), "", E{current_row}*G{current_row}*I{current_row})'
            cell = worksheet.cell(row=current_row, column=10)
            cell.value = formula

            formula = f'=IF(J{current_row}="", "", IF(J{current_row}<10, 1, IF(AND(J{current_row}>=10, J{current_row}<19), 2, IF(AND(J{current_row}>=19, J{current_row}<=27), 3, ""))))'
            cell = worksheet.cell(row=current_row, column=12)
            cell.value = formula

            formula = f'=IF(OR(J{current_row}="", K{current_row}=""), "",J{current_row}*K{current_row})'
            cell = worksheet.cell(row=current_row, column=13)
            cell.value = formula
        
        for idx, col in enumerate(df):
            if idx >= 3:break
            series = df[col]
            max_len = max(
                series.astype(str).map(len).max(),
                len(str(series.name))
            )+5

            current_column_dimension = worksheet.column_dimensions[get_column_letter(idx+1)]
            current_width = current_column_dimension.width

            max_len = max(max_len, current_width)

            worksheet.column_dimensions[get_column_letter(idx+1)].width = max_len