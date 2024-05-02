from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

# 创建一个工作簿
wb = Workbook()

# 创建一个工作表
ws = wb.active

# 设置列名
column_names = ["Column A", "Column B", "Column C"]
ws.append(column_names)

# 创建下拉选单的数据
dropdown_data_A = ["Option 1", "Option 2", "Option 3"]
dropdown_data_B = {"Option 1": ["Option 4", "Option 5"],
                   "Option 2": ["Option 6", "Option 7"],
                   "Option 3": ["Option 8", "Option 9"]}

# 创建下拉选单的数据验证对象
dv_A = DataValidation(type="list", formula1=f'"{",".join(dropdown_data_A)}"', allow_blank=True)
dv_B = DataValidation(type="list", allow_blank=True)

# 将数据验证对象应用到第一列的单元格中
ws.add_data_validation(dv_A)
ws.add_data_validation(dv_B)

for i in range(2, 11):
    dv_A.add(ws[f"A{i}"])  # 第一行是列名，因此数据从第二行开始
    # 检查列 A 的值是否为空
    if ws[f"A{i}"].value is not None and ws[f"A{i}"].value in dropdown_data_B:
        # 根据列 A 的值设置列 B 的下拉选项
        dv_B.formula1 = f'"{",".join(dropdown_data_B[ws[f"A{i}"].value])}"'
        dv_B.add(ws[f"B{i}"])  # 第一行是列名，因此数据从第二行开始

# 保存Excel文件
wb.save("excel_with_dropdown_modified.xlsx")